import openpyxl
import time
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


#--------------------------VARIABLES---------------------------------------------


filename_arrow = '주택용 기초소방시설 보급(만년마을).xlsx'   # 시트 이름 입력
sheet_a_start_row = 4                     # 주소 시작 행 입력
sheet_a_max_row = 43                      # 주소 마지막 행 입력
Fire_EX = 'M'                             # 소화기 열 입력
Fire_SN = 'N'                             # 감지기 열 입력
ADD_col = 'R'                             # 주소 열 입력
Name_col = 'H'                            # 수령자 이름 열 입력
sht_T = 0                                 # 강진 시트 -> 0,  장흥 시트 -> 1 입력
sht_A = 0                                 # 화살 시트 인덱스 -> 첫 번째 시트 일 경우 - 0, 이후 하나씩 증가하여 입력
Year_col = 'O'                            # 연도 열 입력


#---------------------------LOADING-----------------------------------------------


print('Start!')
start = time.time()  # 시작 시간 저장

filename_target = '55강진소방서 주택관리대장.xlsx'


wb_t = openpyxl.load_workbook(filename_target, data_only=True)
sheet_t = wb_t.worksheets[sht_T] #강진시트 혹은 장흥시트

wb_a = openpyxl.load_workbook(filename_arrow, data_only=True)
sheet_a = wb_a.worksheets[sht_A]  #화살 시트 순서

print('load complete!')
print("Loading time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간


#---------------------------WORKING-----------------------------------------------


counter = 0 
Passed = 0
for i in range(sheet_a_start_row, sheet_a_max_row, 1): #''' 바껴야하는 행 ''' #sheet_a.max_row   주소가 몇번째 행부터 시작하는지 꼭 체크하기!
    
    if sheet_a.cell(row= i, column=column_index_from_string(Fire_EX)).value and sheet_a.cell(row= i, column=column_index_from_string(Fire_SN)).value  is None: # '''수령 확인'''  ''' '''
        print('PASS')
        Passed += 1
        continue
    print('NEW CELL DETECTED : arrow I = ' + str(i))
    
    for t in range(6, sheet_t.max_row, 1):  #sheet_t.max_row
        if sheet_a.cell(row=i, column= column_index_from_string(ADD_col)).value == sheet_t.cell(row= t, column= column_index_from_string('BB')).value or sheet_a.cell(row=i, column= column_index_from_string(ADD_col)).value == sheet_t.cell(row= t, column= column_index_from_string('BC')).value:  #주소 일치 확인
            sheet_t.cell(row= t, column= 15).value = 1 #sheet_a.cell(row= i, column= column_index_from_string(Fire_SN)).value  #감지기
            sheet_t.cell(row= t, column= 18).value = 1 #sheet_a.cell(row= i, column= column_index_from_string(Fire_EX)).value  #소화기
            sheet_t.cell(row= t, column= 24).value = sheet_a.cell(row= i, column= column_index_from_string(Name_col)).value  #수령자
            sheet_t.cell(row= t, column= 23).value = sheet_a.cell(row= i, column= column_index_from_string(Year_col)).value  #년도
            sheet_t.cell(row= t, column= 25).value = "안전문화조성"
            counter += 1
            try:
                print(str(counter) + '--' + sheet_a.cell(row= i, column= column_index_from_string(Name_col)).value)
            except TypeError as identifier:
                pass
            
            
print('입력된 셀의 개수 : ' + str(counter))      
print('수령하지 않은 셀의 개수 : ' + str(Passed))
print('Success!!')
print("Work time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
print('NOT DONE YET!!  DO NOT CONTROL PROGRAM!!')


#-----------------------------SAVING---------------------------------------------


wb_t.save('55강진소방서 주택관리대장.xlsx')
print("Save Done!  time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간

print('Job Done!')
print("time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간

