import openpyxl
import time
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


#--------------------------VARIABLES---------------------------------------------

filename_arrow = '기초소방시설 전수조사 결과(장흥센터).xlsx'   # 시트 이름 입력
sheet_a_start_row = 4                     # 주소 시작 행 입력
sheet_a_max_row = 660                    # 주소 마지막 행 입력
Fire_EX = 'J'                             # 소화기 열 입력
Fire_SN = 'K'                             # 감지기 열 입력
Fire_GS = 'L'
ADD_col = 'O'                             # 주소 열 입력
Name_col = 'I'                            # 수령자 이름 열 입력
sht_T = 5                                 # 강진 시트 -> 0,  장흥 시트 -> 1 입력    
sht_A = 1                                 # 화살 시트 인덱스 -> 첫 번째 시트 일 경우 - 0, 이후 하나씩 증가하여 입력
Year_col = 'M'                            # 연도 열 입력


#---------------------------LOADING-----------------------------------------------


print('Start!')
start = time.time()  # 시작 시간 저장

filename_target = '11주택용 소방시설 전수조사(장흥센터).xlsx'


wb_t = openpyxl.load_workbook(filename_target, data_only=True)
sheet_t = wb_t.worksheets[sht_T] #강진시트 혹은 장흥시트

wb_a = openpyxl.load_workbook(filename_arrow, data_only=True)
sheet_a = wb_a.worksheets[sht_A]  #화살 시트 순서

print('load complete!')
print("Loading time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간


#---------------------------WORKING-----------------------------------------------


counter = 0 
Passed = 0
for i in range(sheet_a_start_row, sheet_a_max_row, 1): #''' 바껴야하는 행 ''' #sheet_a.max_row   주소가 몇번째 행부터 시작하는지 꼭 체크하기!
    
    if sheet_a.cell(row= i, column=column_index_from_string(Fire_EX)).value is None and sheet_a.cell(row= i, column=column_index_from_string(Fire_SN)).value is None: # '''수령 확인'''  
        print('PASS')
        Passed += 1
        continue
    print('NEW CELL DETECTED : arrow I = ' + str(i))
    
    for t in range(7, sheet_t.max_row, 1):  #sheet_t.max_row
        # print(sheet_a.cell(row=i, column= column_index_from_string(ADD_col)).value, '----', sheet_t.cell(row= t, column= column_index_from_string('O')).value)
        if sheet_a.cell(row=i, column= column_index_from_string(ADD_col)).value == sheet_t.cell(row= t, column= column_index_from_string('O')).value:  #주소 일치 확인
            sheet_t.cell(row= t, column= 9).value = 1 if sheet_a.cell(row= i, column= column_index_from_string(Fire_GS)).value == '○' else sheet_t.cell(row= t, column= 9).value # 가스차단기
            sheet_t.cell(row= t, column= 7).value = 1 if sheet_a.cell(row= i, column= column_index_from_string(Fire_SN)).value == '○' else sheet_t.cell(row= t, column= 7).value  #감지기
            sheet_t.cell(row= t, column= 8).value = 1 if sheet_a.cell(row= i, column= column_index_from_string(Fire_EX)).value == '○' else sheet_t.cell(row= t, column= 8).value #소화기

            if sheet_a.cell(row= i, column= column_index_from_string(Fire_SN)).value is not None:
                sheet_t.cell(row= t, column= 13).value = sheet_a.cell(row= i, column= column_index_from_string(Name_col)).value  #수령자
                sheet_t.cell(row= t, column= 12).value = 2019 #sheet_a.cell(row= i, column= column_index_from_string(Year_col)).value  #년도
                sheet_t.cell(row= t, column= 11).value = 1 #가구 수
                # sheet_t.cell(row= t, column= 14).value = "화재없는 안전마을"
            counter += 1
            print('ok')
            try:
                print(str(counter) + '--' + sheet_a.cell(row= i, column= column_index_from_string(Name_col)).value)
            except TypeError as identifier:
                pass
            
            
print('입력된 셀의 개수 : ' + str(counter))      
print('수령하지 않은 셀의 개수 : ' + str(Passed))
print('Success!!')
print("Work time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
print('NOT DONE YET!!  DO NOT CONTROL PROGRAM!!')



#-----------------------------SAVING---------------------------------------------

wb_t.save('11주택용 소방시설 전수조사(장흥센터).xlsx')
print("Save Done!  time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간
print('Job Done!')
print("time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간

